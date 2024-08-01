import os.path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.cell.cell import Cell

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import date, timedelta, datetime

import sys
from PyQt5 import QtCore, QtGui, QtWidgets, uic
import PyQt5.QtWidgets

qtCreatorFile = "main.ui"  # Esse é o arquivo .ui gerado pelo QtDesigner
Ui_MainWindow, QtBaseClass = uic.loadUiType(qtCreatorFile)

today = date.today()
workDay = today + timedelta(days=2)
day = workDay.weekday()

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
SAMPLE_SPREADSHEET_ID = (
    "11K0FDN0-8tqhjSUq1Lg8KZmNft0O9hZCtJ4tc0vmu6o"  # ID da planilha google
)


# autenticação do google
def auth():
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())
    return creds


creds = auth()


def main(creds):
    SAMPLE_RANGE_NAME = setRange()
    weekDay = setWeekDay()
    try:
        service = build("sheets", "v4", credentials=creds)

        sheet = service.spreadsheets()
        result = (
            sheet.values()
            .get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME)
            .execute()
        )
        values = result.get("values", [])

        if not values:
            print("Erro ao acessar planilha")
        else:
            for row in values:
                # garantir q n ignora vazios
                if len(row) < 3:
                    row.extend([""] * (3 - len(row)))

            # pega o template
            template_path = "templates/template_" + weekDay + ".xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            # escreve
            for row_idx, row in enumerate(values, start=26):
                for col_idx, value in enumerate(row, start=16):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # corrige férias etc
            lista = []
            file = open("resources/list.txt", "r", encoding="utf-8")
            for line in file:
                name, date, comment = line.strip().split(" - ")
                lista.append((name, date, comment))

            for row in ws:
                for cell in row:
                    if isinstance(cell, Cell) and cell.value is not None:
                        for name, date, comment in lista:
                            returnDate = datetime.strptime(date, "%d/%m/%Y").date()
                            if cell.value == name and workDay < returnDate:
                                comment = Comment(text=comment, author="Author Name")
                                ws.cell(row=cell.row, column=cell.column).comment = (
                                    comment
                                )
                                for i in range(1, 4):
                                    ws.cell(
                                        row=cell.row, column=cell.column + i, value=""
                                    )

            ws.cell(row=3, column=1, value=workDay)
            wb.save("output/ARRANCHAMENTO " + workDay.strftime("%d.%m.%Y") + ".xlsx")

            path = os.path.realpath("output")
            os.startfile(path)

    except HttpError as err:
        print(err)


def setWeekDay():
    if day == 0:
        return "mon"
    elif day < 4:
        return "week"
    else:
        return "wknd"


def setRange():
    if day < 4:
        return "C3:E33"
    elif day == 4:
        return "F3:H33"
    elif day == 5:
        return "I3:K33"
    elif day == 6:
        return "L3:N33"


class MyWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)

        self.pushButton.clicked.connect(self.onClick)

    def onClick(self):

        if today.weekday() != 4:
            main(creds)
        else:
            for i in range(1, 4):
                main(creds)
                workDay += timedelta(days=1)
                day = workDay.weekday()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
